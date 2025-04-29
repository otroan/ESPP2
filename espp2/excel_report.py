"""
Generates Excel reports for the ESPP portfolio.
"""

import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
from espp2 import __version__
from typing import TYPE_CHECKING, List, Tuple, Dict, Any
from decimal import Decimal, ROUND_HALF_UP  # Import Decimal and rounding mode

if TYPE_CHECKING:
    # Avoid circular import, Portfolio needs this module
    from espp2.portfolio import (
        Portfolio,
        PortfolioPosition,
        PortfolioDividend,
        PortfolioSale,
        PortfolioTransfer,
    )
    # Type hints for data structures used
    # from espp2.positions import Ledger # If needed

# --- Module-level runtime import (Restored) ---
from espp2.portfolio import (
    PortfolioPosition,
    PortfolioDividend,
    PortfolioSale,
    PortfolioTransfer,
)

logger = logging.getLogger(__name__)

# --- Constants for Formatting ---
CURRENCY_FORMAT = "0.00"  # Format for currency values
QTY_FORMAT = "0.0000"  # Format for quantity values
CURRENCY_QUANTIZER = Decimal("0.01")  # Quantizer for currency
QTY_QUANTIZER = Decimal("0.0001")  # Quantizer for quantity

# --- Helper Functions ---


def format_cells(ws, column_letter: str, number_format: str):
    """Sets the number format for all cells in a given column (skipping header row 1 & 2)."""
    # Assumes headers are in row 1/2, data starts row 3.
    for row in range(3, ws.max_row + 1):
        cell_ref = f"{column_letter}{row}"
        try:  # Protect against potential errors on merged/empty cells
            cell = ws[cell_ref]
            cell.number_format = number_format

        except AttributeError:
            logger.debug(f"Could not format cell {cell_ref}")
        # Handle potential KeyError if cell doesn't exist (less likely but possible)
        except KeyError:
            logger.debug(f"Cell {cell_ref} does not exist, skipping format.")


def adjust_width(ws):
    """Adjusts column width to fit the longest value in each column, including header."""

    def as_text(value):
        if value is None:
            return ""
        # Handle date objects specifically for length calculation
        if isinstance(value, date):
            return value.isoformat()  # YYYY-MM-DD format
        return str(value)

    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                cell_value_str = as_text(cell.value)
                # Handle potential formulas by checking calculated value if available
                # This part might need refinement depending on openpyxl version/behavior
                # if cell.data_type == 'f':
                #    # Attempt to get calculated value length, fallback to formula length
                #    try: cell_value_str = as_text(cell.internal_value) # Or check cell._value
                #    except: pass
                cell_len = len(cell_value_str)
                if cell_len > max_length:
                    max_length = cell_len
            except Exception as e:
                logger.debug(f"Could not get length for cell {column_letter}{row}: {e}")
                pass

        # --- REDUCE PADDING for narrower columns ---
        # adjusted_width = max_length + 0.5 # Reduced padding from 1.5
        # if adjusted_width < 8: # Minimum width (keep for now)
        #     adjusted_width = 8
        # ws.column_dimensions[column_letter].width = adjusted_width


def index_to_cell(row: int, column_index: int) -> str:
    """
    Convert a 1-based row and 0-based column index to an Excel cell reference.
    """
    if column_index < 0:
        raise ValueError("Column index cannot be negative")
    column_letter = get_column_letter(column_index + 1)
    return f"{column_letter}{row}"


# --- Row Formatting Functions (New) ---


def format_position_row(
    position: PortfolioPosition, row: int, header_map: Dict[str, int]
) -> Tuple[List[Tuple[int, int, Any]], Dict[str, str]]:
    """Formats the data for a PortfolioPosition row and returns cell tuples and coordinates."""
    cells = []
    coords = {}
    col_indices = {name: idx for name, idx in header_map.items()}  # Easier access

    cells.append((row, col_indices["Symbol"], position.symbol))
    if not position.split:
        cells.append((row, col_indices["Date"], position.date))
    if position.pre_split_qty > 0:
        cells.append((row, col_indices["pQty"], position.pre_split_qty))
    cells.append((row, col_indices["Qty"], position.qty))

    # Price Calculation
    price_usd_idx = col_indices["Price USD"]
    exchange_rate_idx = col_indices["Exch. Rate"]
    price_idx = col_indices["Price"]
    price_usd_cell = index_to_cell(row, price_usd_idx)
    exchange_rate_cell = index_to_cell(row, exchange_rate_idx)
    price_cell = index_to_cell(row, price_idx)
    cells.append((row, price_idx, f"={price_usd_cell}*{exchange_rate_cell}"))
    coords["Price"] = price_cell
    coords["Price USD"] = price_usd_cell

    # Quantize values before adding, ensuring they are Decimal first
    purchase_price_val = position.purchase_price.value.quantize(
        CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP
    )

    tax_acc_val = position.tax_deduction_acc
    if not isinstance(tax_acc_val, Decimal):
        tax_acc_val = Decimal(str(tax_acc_val))  # Convert int/float
    tax_acc = tax_acc_val.quantize(CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP)

    tax_new_val = position.tax_deduction_new
    if not isinstance(tax_new_val, Decimal):
        tax_new_val = Decimal(str(tax_new_val))  # Convert int/float
    tax_new = tax_new_val.quantize(CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP)

    exchange_rate = (
        position.purchase_price.nok_exchange_rate
    )  # Exchange rates might need specific precision?
    if exchange_rate:  # Assuming Decimal
        # Let's use more precision for exchange rates typically
        exchange_rate = exchange_rate.quantize(
            Decimal("0.0001"), rounding=ROUND_HALF_UP
        )

    cells.append((row, price_usd_idx, purchase_price_val))
    cells.append((row, exchange_rate_idx, exchange_rate))
    cells.append((row, col_indices["Acc. TxtDed"], tax_acc))
    cells.append((row, col_indices["New TxtDed"], tax_new))

    return cells, coords


def format_dividend_row(
    dividend: PortfolioDividend, row: int, header_map: Dict[str, int]
) -> List[Tuple[int, int, Any]]:
    """Formats the data for a PortfolioDividend row."""
    cells = []
    col_indices = {name: idx for name, idx in header_map.items()}

    cells.append((row, col_indices["Date"], dividend.divdate))
    cells.append((row, col_indices["Type"], "Dividend"))
    cells.append((row, col_indices["iQty"], dividend.qty))

    exchange_rate_idx = col_indices["Exch. Rate"]
    div_ps_usd_idx = col_indices["Div/Share USD"]
    div_ps_idx = col_indices["Div/Share"]
    iqty_idx = col_indices["iQty"]
    total_div_idx = col_indices["Tot. Div."]
    total_div_usd_idx = col_indices["Tot. Div. USD"]

    exchange_rate_cell = index_to_cell(row, exchange_rate_idx)
    div_ps_usd_cell = index_to_cell(row, div_ps_usd_idx)
    div_ps_cell = index_to_cell(row, div_ps_idx)
    iqty_cell = index_to_cell(row, iqty_idx)
    # total_div_cell = index_to_cell(row, total_div_idx)
    # total_div_usd_cell = index_to_cell(row, total_div_usd_idx)

    # Quantize values, ensuring Decimal type
    div_qty = dividend.qty.quantize(QTY_QUANTIZER, rounding=ROUND_HALF_UP)
    exchange_rate = dividend.dividend_dps.nok_exchange_rate
    if exchange_rate:
        exchange_rate = exchange_rate.quantize(
            Decimal("0.0001"), rounding=ROUND_HALF_UP
        )
    div_dps_val = dividend.dividend_dps.value.quantize(
        CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP
    )

    tax_used_val = dividend.tax_deduction_used
    if not isinstance(tax_used_val, Decimal):
        tax_used_val = Decimal(str(tax_used_val))
    tax_used = tax_used_val.quantize(CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP)

    tax_total_val = dividend.tax_deduction_used_total
    if not isinstance(tax_total_val, Decimal):
        tax_total_val = Decimal(str(tax_total_val))
    tax_total = tax_total_val.quantize(CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP)

    cells.append((row, col_indices["iQty"], div_qty))  # Use quantized qty
    cells.append((row, exchange_rate_idx, exchange_rate))
    cells.append((row, div_ps_idx, f"={div_ps_usd_cell}*{exchange_rate_cell}"))
    cells.append((row, div_ps_usd_idx, div_dps_val))
    cells.append((row, total_div_idx, f"={div_ps_cell}*{iqty_cell}"))
    cells.append((row, total_div_usd_idx, f"={div_ps_usd_cell}*{iqty_cell}"))
    cells.append((row, col_indices["Used TxtDed"], tax_used))
    cells.append((row, col_indices["Rem. TxtDed"], tax_total))

    return cells


def format_sale_row(
    sale: PortfolioSale,
    row: int,
    header_map: Dict[str, int],
    parent_coords: Dict[str, str],
) -> List[Tuple[int, int, Any]]:
    """Formats the data for a PortfolioSale row, using parent coordinates."""
    cells = []
    col_indices = {name: idx for name, idx in header_map.items()}

    parent_price_cell = parent_coords.get("Price", "#REF!")  # Get parent coords
    parent_price_usd_cell = parent_coords.get("Price USD", "#REF!")

    cells.append((row, col_indices["Date"], sale.saledate))
    cells.append((row, col_indices["Type"], "Sale"))
    cells.append((row, col_indices["Qty"], sale.qty))

    price_usd_idx = col_indices["Price USD"]
    exchange_rate_idx = col_indices["Exch. Rate"]
    price_idx = col_indices["Price"]
    qty_idx = col_indices["Qty"]
    gain_ps_idx = col_indices["Gain/Share"]
    gain_ps_usd_idx = col_indices["Gain/Share USD"]
    gain_idx = col_indices["Gain"]
    gain_usd_idx = col_indices["Gain USD"]
    amount_idx = col_indices["Amount"]
    amount_usd_idx = col_indices["Amt USD"]

    price_usd_cell = index_to_cell(row, price_usd_idx)
    exchange_rate_cell = index_to_cell(row, exchange_rate_idx)
    price_cell = index_to_cell(row, price_idx)
    qty_cell = index_to_cell(row, qty_idx)
    gain_ps_cell = index_to_cell(row, gain_ps_idx)
    gain_ps_usd_cell = index_to_cell(row, gain_ps_usd_idx)
    #     gain_cell = index_to_cell(row, gain_idx)
    #     gain_usd_cell = index_to_cell(row, gain_usd_idx)
    #     amount_cell = index_to_cell(row, amount_idx)
    #     amount_usd_cell = index_to_cell(row, amount_usd_idx)

    # Quantize values, ensuring Decimal type
    sale_qty = sale.qty.quantize(QTY_QUANTIZER, rounding=ROUND_HALF_UP)
    sell_price_val = sale.sell_price.value.quantize(
        CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP
    )
    exchange_rate = sale.sell_price.nok_exchange_rate
    if exchange_rate:
        exchange_rate = exchange_rate.quantize(
            Decimal("0.0001"), rounding=ROUND_HALF_UP
        )
    tax_used_val = sale.tax_deduction_used
    if not isinstance(tax_used_val, Decimal):
        tax_used_val = Decimal(str(tax_used_val))
    tax_used = tax_used_val.quantize(CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP)
    tax_total_val = sale.tax_deduction_used_total
    if not isinstance(tax_total_val, Decimal):
        tax_total_val = Decimal(str(tax_total_val))
    tax_total = tax_total_val.quantize(CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP)

    cells.append((row, col_indices["Qty"], sale_qty))  # Use quantized qty
    cells.append((row, price_idx, f"={price_usd_cell}*{exchange_rate_cell}"))
    cells.append((row, price_usd_idx, sell_price_val))
    cells.append((row, exchange_rate_idx, exchange_rate))

    cells.append((row, gain_ps_idx, f"={price_cell}-{parent_price_cell}"))
    cells.append((row, gain_ps_usd_idx, f"={price_usd_cell}-{parent_price_usd_cell}"))

    # RESTORED formulas for Gain / Gain USD
    cells.append((row, gain_idx, f"={gain_ps_cell}*ABS({qty_cell})"))
    cells.append((row, gain_usd_idx, f"={gain_ps_usd_cell}*ABS({qty_cell})"))

    cells.append((row, amount_idx, f"=ABS({price_cell}*{qty_cell})"))
    cells.append((row, amount_usd_idx, f"=ABS({price_usd_cell}*{qty_cell})"))
    cells.append((row, col_indices["Used TxtDed"], tax_used))
    cells.append((row, col_indices["Rem. TxtDed"], tax_total))

    return cells


def format_transfer_row(
    transfer: PortfolioTransfer, row: int, header_map: Dict[str, int]
) -> List[Tuple[int, int, Any]]:
    """Formats the data for a PortfolioTransfer row."""
    cells = []
    col_indices = {name: idx for name, idx in header_map.items()}

    # Quantize qty
    transfer_qty = transfer.qty.quantize(QTY_QUANTIZER, rounding=ROUND_HALF_UP)

    cells.append((row, col_indices["Date"], transfer.date))
    cells.append((row, col_indices["Type"], "Transfer"))
    cells.append((row, col_indices["Qty"], transfer_qty))

    return cells


# --- Main Excel Report Generation Function (Modified) ---


def generate_workbook(portfolio: "Portfolio") -> BytesIO:
    """Generates the full Excel workbook for the portfolio."""
    year = portfolio.year
    positions = portfolio.positions
    column_headers = portfolio.column_headers
    workbook = Workbook()

    # --- Portfolio Sheet ---
    ws = workbook.active
    ws.title = f"Portfolio-{year}"
    disclaimer = (
        "Disclaimer: This tool is provided as is, without warranty of any kind. "
        "Use of this tool is at your own risk. The authors or distributors "
        "are not responsible for any losses, damages, or issues that may arise "
        "from using this tool. Always consult with a professional financial advisor "
        "before making any financial decisions. "
        f"This report is generated with the espp2 tool version: {__version__} on {date.today().isoformat()}"
    )

    # Merged Title Rows (Row 1)
    ws.merge_cells("J1:M1")  # Dividends title span
    ws["J1"] = "Dividends"
    ws["J1"].font = Font(bold=True)
    ws["J1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells("N1:Q1")  # Tax Deduction title span
    ws["N1"] = "Deductible Risk-free return"
    ws["N1"].font = Font(bold=True)
    ws["N1"].alignment = Alignment(horizontal="center", vertical="center")
    # Assuming 'Sales' columns might be R through W (adjust if needed)
    last_sales_col = (
        get_column_letter(column_headers.index("Amount USD") + 1)
        if "Amount USD" in column_headers
        else "W"
    )
    ws.merge_cells(f"R1:{last_sales_col}1")
    ws["R1"] = "Sales"
    ws["R1"].font = Font(bold=True)
    ws["R1"].alignment = Alignment(horizontal="center", vertical="center")

    # Column Headers (Row 2)
    ws.append(column_headers)  # Appends to row 2
    header_row_idx = 2
    for cell in ws[header_row_idx]:
        cell.font = Font(bold=True)

    # Header to Index Map
    header_to_index = {header: i for i, header in enumerate(column_headers)}

    # Write data using new formatting functions
    data_start_row = 3
    current_row = data_start_row
    position_coords_map = {}  # Store coordinates for parent lookup

    for pos_idx, stock_position in enumerate(positions):
        if not isinstance(stock_position, PortfolioPosition):
            logger.warning(
                f"Skipping unexpected item in positions list: {type(stock_position)}"
            )
            continue

        # Format position row and get coordinates
        position_cells, coords = format_position_row(
            stock_position, current_row, header_to_index
        )
        position_coords_map[pos_idx] = coords  # Use index as key

        # Write position cells
        for r, col_idx, value in position_cells:
            ws.cell(row=r, column=col_idx + 1, value=value)
        current_row += 1

        for record in stock_position.records:
            record_cells = []
            parent_coords = None
            # Find the index of the parent position to use as the map key
            try:
                parent_idx = positions.index(record.parent)
                parent_coords = position_coords_map.get(
                    parent_idx
                )  # Get parent coords using index
            except ValueError:  # Handle case where parent might not be in the list (shouldn't happen ideally)
                logger.error(
                    f"Parent position for record at row {current_row} not found in positions list."
                )
            except AttributeError:  # Handle if record has no parent
                logger.warning(f"Record at row {current_row} has no parent attribute.")

            if isinstance(record, PortfolioSale):
                if parent_coords:
                    record_cells = format_sale_row(
                        record, current_row, header_to_index, parent_coords
                    )
                else:
                    logger.warning(
                        f"Could not find parent coords for sale record at row {current_row}"
                    )
                    # Optionally format with '#REF!' or skip
            elif isinstance(record, PortfolioDividend):
                # Dividends don't strictly need parent coords for their own formatting
                record_cells = format_dividend_row(record, current_row, header_to_index)
            elif isinstance(record, PortfolioTransfer):
                record_cells = format_transfer_row(record, current_row, header_to_index)
            else:
                logger.debug(
                    f"Skipping formatting for unknown record type: {type(record)}"
                )

            # Write record cells
            for r, col_idx, value in record_cells:
                ws.cell(row=r, column=col_idx + 1, value=value)
            current_row += 1  # Move to next row

    # Create header to column letter mapping (needed for formatting functions below)
    header_to_letter = {
        header: get_column_letter(i + 1) for i, header in enumerate(column_headers)
    }

    # --- Formatting Portfolio Sheet ---
    num_columns_2dp = [
        "Price",
        "Price USD",
        "Gain",
        "Gain/Share",
        "Gain USD",
        "Amount",
        "Amt USD",
        "Div/Share",
        "Div/Share USD",
        "Tot. Div.",
        "Tot. Div. USD",
        "Exch. Rate",
        "Acc. TxtDed",
        "New TxtDed",
        "Used TxtDed",
        "Rem. TxtDed",
    ]
    num_cols_2dp_letters = [
        header_to_letter[h] for h in num_columns_2dp if h in header_to_letter
    ]
    for col_letter in num_cols_2dp_letters:
        # Use constant for currency format
        format_cells(ws, col_letter, CURRENCY_FORMAT)

    num_columns_4dp = ["pQty", "Qty", "iQty"]
    num_cols_4dp_letters = [
        header_to_letter[h] for h in num_columns_4dp if h in header_to_letter
    ]
    for col_letter in num_cols_4dp_letters:
        # Use constant for quantity format
        format_cells(ws, col_letter, QTY_FORMAT)

    # Freeze Panes (freeze rows 1 and 2)
    ws.freeze_panes = ws["A3"]

    # Sum Totals row (calculate index based on current_row)
    total_row_idx = current_row  # The row after the last data row
    sum_columns = [
        # "Qty", # Summing current Qty might not be meaningful if there are splits/transfers
        "Gain",
        "Gain USD",
        "Amount",
        "Amt USD",
        "Tot. Div.",
        "Tot. Div. USD",
        "Rem. TxtDed",
    ]
    sum_cols_letters = [
        header_to_letter[h] for h in sum_columns if h in header_to_letter
    ]

    ws[f"A{total_row_idx}"] = "Total"
    ws[f"A{total_row_idx}"].font = Font(bold=True)

    for col_letter in sum_cols_letters:
        # Sum from data start row up to the last data row (total_row_idx - 1)
        formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row_idx - 1})"
        cell = ws[f"{col_letter}{total_row_idx}"]
        cell.value = formula
        cell.font = Font(bold=True)
        # Use constant for currency format in totals
        cell.number_format = CURRENCY_FORMAT

    # --- Apply Fill Colors to Whole Columns (Header to Total Row) ---
    fill_dividend = PatternFill(
        start_color="CAD8EE", end_color="CAD8EE", fill_type="solid"
    )
    fill_gain = PatternFill(start_color="90ADD7", end_color="90ADD7", fill_type="solid")
    fill_taxded = PatternFill(
        start_color="618CCE", end_color="618CCE", fill_type="solid"
    )

    dividend_headers = ["Div/Share", "Div/Share USD", "Tot. Div.", "Tot. Div. USD"]
    gain_headers = [
        "Gain/Share",
        "Gain/Share USD",
        "Gain",
        "Gain USD",
        "Amount",
        "Amt USD",
    ]
    taxded_headers = ["Acc. TxtDed", "New TxtDed", "Rem. TxtDed", "Used TxtDed"]

    # Apply Dividend Fill
    for header in dividend_headers:
        if header in header_to_letter:
            col_letter = header_to_letter[header]
            for row in range(header_row_idx, total_row_idx + 1):
                ws[f"{col_letter}{row}"].fill = fill_dividend

    # Apply Gain/Sale Fill
    for header in gain_headers:
        if header in header_to_letter:
            col_letter = header_to_letter[header]
            for row in range(header_row_idx, total_row_idx + 1):
                ws[f"{col_letter}{row}"].fill = fill_gain

    # Apply Tax Deduction Fill
    for header in taxded_headers:
        if header in header_to_letter:
            col_letter = header_to_letter[header]
            for row in range(header_row_idx, total_row_idx + 1):
                ws[f"{col_letter}{row}"].fill = fill_taxded

    # Apply conditional formatting for negative numbers (red font)
    # Apply to the data range, excluding totals row
    data_range = (
        f"A{data_start_row}:{get_column_letter(len(column_headers))}{total_row_idx - 1}"
    )
    # Ensure rule doesn't stop other rules if needed later
    ws.conditional_formatting.add(
        data_range,
        CellIsRule(
            operator="lessThan",
            formula=["0"],
            stopIfTrue=False,
            font=Font(color="00FF0000"),
        ),
    )

    # adjust_width(ws) # Keep commented out
    # Manually set Date column width
    ws.column_dimensions["B"].width = 11  # Portfolio Sheet Date (Column B)
    # Manually set Amount column widths
    if "Amount" in header_to_letter:
        ws.column_dimensions[header_to_letter["Amount"]].width = 14
    if "Amt USD" in header_to_letter:
        ws.column_dimensions[header_to_letter["Amt USD"]].width = 14

    # Write the disclaimer below the totals
    disclaimer_row_idx = total_row_idx + 4  # Add some space
    ws[f"A{disclaimer_row_idx}"] = disclaimer
    ws[f"A{disclaimer_row_idx}"].alignment = Alignment(wrapText=True)
    # Optional: Merge cells for the disclaimer to make it span wider
    #     end_col_letter = get_column_letter(
    #     min(5, len(column_headers))
    # )  # Span first 5 cols or less
    ws.merge_cells(
        start_row=disclaimer_row_idx,
        start_column=1,
        end_row=disclaimer_row_idx + 2,
        end_column=min(5, len(column_headers)),
    )

    # --- Cash Sheet ---
    ws_cash = workbook.create_sheet("Cash")
    cash_headers = [
        "Date",
        "Description",
        "Amount NOK",
        "Amount Base",
        "Currency",
        "Balance NOK",
    ]
    ws_cash.append(cash_headers)
    ws_cash.freeze_panes = ws_cash["A2"]
    for cell in ws_cash[1]:  # Bold headers
        cell.font = Font(bold=True)

    # Assuming portfolio.cash_ledger is List[Tuple[CashEntry, Decimal]]
    for entry, balance_nok in portfolio.cash_ledger:
        # Quantize values before appending
        nok_value_quantized = None
        if entry.amount.nok_value is not None:
            try:
                nok_value_quantized = entry.amount.nok_value.quantize(
                    CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP
                )
            except AttributeError:  # Handle if nok_value is not Decimal
                nok_value_quantized = entry.amount.nok_value  # Keep original

        amount_base_quantized = None
        if entry.amount.value is not None:
            try:
                amount_base_quantized = entry.amount.value.quantize(
                    CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP
                )
            except AttributeError:  # Handle if value is not Decimal
                amount_base_quantized = entry.amount.value  # Keep original

        balance_nok_quantized = None
        if balance_nok is not None:
            try:
                balance_nok_quantized = balance_nok.quantize(
                    CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP
                )
            except AttributeError:
                balance_nok_quantized = balance_nok

        ws_cash.append(
            [
                entry.date,
                entry.description,
                nok_value_quantized if nok_value_quantized is not None else "N/A",
                amount_base_quantized if amount_base_quantized is not None else "N/A",
                entry.amount.currency,
                balance_nok_quantized if balance_nok_quantized is not None else "N/A",
            ]
        )

    # Formatting for Cash sheet
    # Use constant for currency format
    format_cells(ws_cash, "C", CURRENCY_FORMAT)  # Amount NOK
    format_cells(ws_cash, "D", CURRENCY_FORMAT)  # Amount Base
    format_cells(ws_cash, "F", CURRENCY_FORMAT)  # Balance NOK
    # adjust_width(ws_cash) # Keep commented out
    # Manually set Date column width
    ws_cash.column_dimensions["A"].width = 11  # Cash Sheet Date (Column A)
    # Manually set Amount column widths
    ws_cash.column_dimensions["C"].width = 14  # Amount NOK (Column C)
    ws_cash.column_dimensions["D"].width = 14  # Amount Base (Column D)

    # --- EOY Holdings Sheet ---
    ws_eoy = workbook.create_sheet("EOY Holdings")
    eoy_headers = [
        "Symbol",
        "Purchase Date",
        "Qty",
        "Purchase Price NOK",
        "Available Tax Deduction (Skjerming) NOK",
    ]
    ws_eoy.append(eoy_headers)
    ws_eoy.freeze_panes = ws_eoy["A2"]
    for cell in ws_eoy[1]:  # Bold headers
        cell.font = Font(bold=True)

    # Assuming portfolio.eoy_holdings is Holdings model
    if portfolio.eoy_holdings and portfolio.eoy_holdings.stocks:
        for h in portfolio.eoy_holdings.stocks:
            # Need to ensure purchase_price Amount object exists and has nok_value calculated
            purchase_nok = None
            try:
                purchase_nok = h.purchase_price.nok_value
            except Exception as e:  # Catch potential errors if nok_value isn't computed
                logger.warning(
                    f"Could not get NOK purchase price for holding {h.symbol} {h.date}: {e}"
                )

            # Quantize values before appending
            qty_quantized = None
            if h.qty is not None:
                try:
                    qty_quantized = h.qty.quantize(
                        QTY_QUANTIZER, rounding=ROUND_HALF_UP
                    )
                except AttributeError:
                    qty_quantized = h.qty

            purchase_nok_quantized = None
            if purchase_nok is not None:
                try:
                    purchase_nok_quantized = purchase_nok.quantize(
                        CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP
                    )
                except AttributeError:
                    purchase_nok_quantized = purchase_nok

            tax_deduction_quantized = None
            if h.tax_deduction is not None:
                try:
                    tax_deduction_quantized = h.tax_deduction.quantize(
                        CURRENCY_QUANTIZER, rounding=ROUND_HALF_UP
                    )
                except AttributeError:
                    tax_deduction_quantized = h.tax_deduction

            ws_eoy.append(
                [
                    h.symbol,
                    h.date,
                    qty_quantized
                    if qty_quantized is not None
                    else "N/A",  # Use quantized
                    purchase_nok_quantized
                    if purchase_nok_quantized is not None
                    else "N/A",  # Use quantized
                    tax_deduction_quantized
                    if tax_deduction_quantized is not None
                    else "N/A",  # Use quantized
                ]
            )

    # Formatting for EOY Holdings sheet
    # Use constants
    format_cells(ws_eoy, "C", QTY_FORMAT)  # Qty
    format_cells(ws_eoy, "D", CURRENCY_FORMAT)  # Purchase Price NOK
    format_cells(ws_eoy, "E", CURRENCY_FORMAT)  # Tax Deduction NOK
    # adjust_width(ws_eoy) # Keep commented out
    # Manually set Date column width
    ws_eoy.column_dimensions["B"].width = 11  # EOY Holdings Purchase Date (Column B)
    # Manually set Amount column width
    ws_eoy.column_dimensions["D"].width = 14  # Purchase Price NOK (Column D)

    # --- Save workbook to buffer ---
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
