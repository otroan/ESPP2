"""
ESPP portfolio class
"""
import logging
from io import BytesIO
from copy import deepcopy
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, field_validator
from typing import Optional
from datetime import date, datetime
from decimal import Decimal
from espp2.datamodels import (
    Stock,
    Holdings,
    Amount,
    Wires,
    Transactions,
    Dividend_Reinv,
    EOYBalanceItem,
    TaxSummary,
    ForeignShares,
    EOYDividend,
    EOYSales,
    SalesPosition,
    Dividend,
    Tax,
    PositiveAmount,
    NegativeAmount,
    GainAmount,
    NativeAmount,
)
from espp2.fmv import FMV, get_tax_deduction_rate, Fundamentals, todate
from espp2.cash import Cash
from espp2.positions import Ledger
from typing import Any, Dict
from espp2.report import print_cash_ledger
from espp2.console import console
from espp2.util import FeatureFlagEnum

fmv = FMV()
logger = logging.getLogger(__name__)

# Temporarily hard-code version
version='0.1.dev300+g328e721.d20240417'

def format_cells(ws, column, number_format):
    for cell in ws[column]:
        cell.number_format = number_format

def format_fill_columns(ws, headers, columns, color):
    # Create a dictionary mapping column headers to Excel column letters
    header_to_letter = {header: chr(i + 65) for i, header in enumerate(headers)}
    cols = [header_to_letter[header] for header in columns if header in header_to_letter]
    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for c in cols:
        for cell in ws[c]:
            cell.fill = fill


class PortfolioPosition(BaseModel):
    """Stock positions"""

    symbol: str
    date: date
    qty: Decimal
    pre_split_qty: Decimal = 0
    tax_deduction: Decimal = 0  # Total available tax deduction
    tax_deduction_acc: Decimal  # accumulated tax deduction from previous years
    tax_deduction_new: Decimal = 0  # tax deduction for this year
    purchase_price: Amount
    discounted_purchase_price: Optional[Amount] = None
    current_qty: Decimal = 0
    records: list[Any] = []
    coord: Dict[str, str] = {}
    split: bool = False

    @field_validator('pre_split_qty', mode='before')
    @classmethod
    def set_pre_split_qty(cls, v, info):
        return v or info.data.get('qty')

    def get_coord(self, key):
        return self.coord[key]

    def qty_at_date(self, exdate):
        """Return qty at date"""
        if self.date > exdate:
            return 0
        qty = self.qty
        for r in self.records:
            if isinstance(r, (PortfolioSale, PortfolioTransfer)):
                transdate = r.saledate if isinstance(r, PortfolioSale) else r.date
                if transdate < exdate:
                    qty -= abs(r.qty)
        return qty

    def format(self, row, columns):
        """Return a list of cells for a row"""
        col = [(row, columns.index("Symbol"), self.symbol)]
        if not self.split:
            col.append((row, columns.index("Date"), self.date))
        if self.pre_split_qty > 0:
            col.append((row, columns.index("pQty"), self.pre_split_qty))
        col.append((row, columns.index("Qty"), self.qty))
        col.append(
            (
                row,
                columns.index("Price"),
                f'={index_to_cell(row, columns.index("Price USD"))}*{index_to_cell(row, columns.index("Exchange Rate"))}',
            )
        )
        self.coord["Price"] = index_to_cell(row, columns.index("Price"))
        self.coord["Price USD"] = index_to_cell(row, columns.index("Price USD"))
        col.append((row, columns.index("Price USD"), round(self.purchase_price.value, 2)))
        col.append(
            (row, columns.index("Exchange Rate"), self.purchase_price.nok_exchange_rate)
        )
        col.append((row, columns.index("Accumulated"), round(self.tax_deduction_acc, 2)))
        col.append((row, columns.index("Added"), round(self.tax_deduction_new, 2)))
        return col


class PortfolioDividend(BaseModel):
    """Stock dividends"""

    divdate: date
    qty: Decimal
    dividend_dps: Amount
    dividend: Amount
    tax_deduction_used: Decimal = 0
    tax_deduction_used_total: Decimal = 0
    parent: PortfolioPosition = None

    def format(self, row, columns):
        """Return a list of cells for a row"""
        col = [(row, columns.index("Date"), self.divdate)]
        col.append((row, columns.index("Type"), "Dividend"))
        col.append((row, columns.index("iQty"), self.qty))
        col.append(
            (row, columns.index("Exchange Rate"), self.dividend_dps.nok_exchange_rate)
        )
        col.append(
            (
                row,
                columns.index("Div PS"),
                f'={index_to_cell(row, columns.index("Div PS USD"))}*{index_to_cell(row, columns.index("Exchange Rate"))}',
            )
        )
        col.append((row, columns.index("Div PS USD"), round(self.dividend_dps.value, 2)))
        col.append(
            (
                row,
                columns.index("Total Dividend"),
                f'={index_to_cell(row, columns.index("Div PS"))}*{index_to_cell(row, columns.index("iQty"))}',
            )
        )

        col.append(
            (
                row,
                columns.index("Total Dividend USD"),
                f'={index_to_cell(row, columns.index("Div PS USD"))}*{index_to_cell(row, columns.index("iQty"))}',
            )
        )
        col.append(
            (row, columns.index("Used"), round(self.tax_deduction_used, 2))
        )
        col.append(
            (
                row,
                columns.index("TD Total"),
                round(self.tax_deduction_used_total, 2),
            )
        )
        return col


class PortfolioSale(BaseModel):
    # TODO: Fee
    saledate: date
    qty: Decimal
    sell_price: Amount
    gain_ps: GainAmount
    gain: GainAmount
    total: Amount
    tax_deduction_used: Decimal = 0
    tax_deduction_used_total: Decimal = 0
    parent: PortfolioPosition = None
    id: str

    def format(self, row, columns):
        col = [(row, columns.index("Date"), self.saledate)]
        col.append((row, columns.index("Type"), "Sale"))
        col.append((row, columns.index("Qty"), self.qty))
        col.append(
            (
                row,
                columns.index("Price"),
                f'={index_to_cell(row, columns.index("Price USD"))}*{index_to_cell(row, columns.index("Exchange Rate"))}',
            )
        )
        col.append((row, columns.index("Price USD"), self.sell_price.value))
        col.append(
            (row, columns.index("Exchange Rate"), self.sell_price.nok_exchange_rate)
        )
        col.append(
            (
                row,
                columns.index("Gain PS"),
                f'={index_to_cell(row, columns.index("Price"))}-{self.parent.get_coord("Price")}',
            )
        )

        col.append(
            (
                row,
                columns.index("Gain PS USD"),
                f'={index_to_cell(row, columns.index("Price USD"))}-{self.parent.get_coord("Price USD")}',
            )
        )

        col.append(
            (
                row,
                columns.index("Gain"),
                f'={index_to_cell(row, columns.index("Gain PS"))}*ABS({index_to_cell(row, columns.index("Qty"))})',
            )
        )
        col.append(
            (
                row,
                columns.index("Gain USD"),
                f'={index_to_cell(row, columns.index("Gain PS USD"))}*ABS({index_to_cell(row, columns.index("Qty"))})',
            )
        )
        col.append(
            (
                row,
                columns.index("Amount"),
                f'=ABS({index_to_cell(row, columns.index("Price"))}*{index_to_cell(row, columns.index("Qty"))})',
            )
        )
        col.append(
            (
                row,
                columns.index("Amount USD"),
                f'=ABS({index_to_cell(row, columns.index("Price USD"))}*{index_to_cell(row, columns.index("Qty"))})',
            )
        )
        col.append(
            (row, columns.index("Used"), round(self.tax_deduction_used, 2))
        )
        col.append(
            (
                row,
                columns.index("TD Total"),
                round(self.tax_deduction_used_total, 2),
            )
        )
        return col

class PortfolioTransfer(BaseModel):
    date: date
    qty: Decimal
    parent: PortfolioPosition = None
    id: str

    def format(self, row, columns):
        col = [(row, columns.index("Date"), self.date)]
        col.append((row, columns.index("Type"), "Transfer"))
        col.append((row, columns.index("Qty"), self.qty))
        return col

def adjust_width(ws):
    def as_text(value):
        if value is None:
            return ""
        return str(value)

    # Adjust column width to fit the longest value in each column
    for column_cells in ws.columns:
        column_cells = column_cells[1:]
        if not column_cells:
            continue
        length = max(len(as_text(cell.value)) for cell in column_cells)
        if length < 8:
            length = 8
        try:
            ws.column_dimensions[column_cells[0].column_letter].width = length
        except:
            pass


def index_to_cell(row, column):
    """
    Convert a row and column index to an Excel cell reference.
    """
    column_letter = get_column_letter(column + 1)
    return f"{column_letter}{row}"


class Portfolio:
    def buy(self, p):
        position = PortfolioPosition(
                qty=p.qty,
                purchase_price=p.purchase_price,
                discounted_purchase_price=p.discounted_purchase_price,
                date=p.date,
                symbol=p.symbol,
                tax_deduction_acc=0,
                current_qty=p.qty,
            )
        self.positions.append(position)

        # Keep stock of new positions for reporting
        self.new_positions.append(position)

    def qty_at_date(self, symbol, exdate):
        total = Decimal('0.00')
        for p in self.positions:
            if p.symbol == symbol:
                total += p.qty_at_date(exdate)
        return total

    def dividend(self, transaction):
        """Dividend"""
        shares_left = transaction.amount.value / transaction.dividend_dps
        expected_number_of_shares = shares_left
        found_number_of_shares = 0
        total = transaction.amount.value
        # Walk through positions available at exdate.
        no_shares = self.qty_at_date(transaction.symbol, transaction.exdate)
        expected_dividend = round(no_shares * transaction.dividend_dps, 2)
        if expected_dividend != transaction.amount.value:
            logger.error(f"Dividend error. {transaction.date} Expected {expected_number_of_shares} shares, holding: {no_shares}")
        for p in self.positions:
            if p.symbol == transaction.symbol:
                # Get qty up until exdate
                qty = p.qty_at_date(transaction.exdate)
                if qty == 0:
                    continue
                assert (
                    p.date <= transaction.exdate
                ), f"Exdate {transaction.exdate} before purchase date {p.date} shares left {shares_left}"
                if qty >= shares_left:
                    shares_left = 0
                else:
                    shares_left -= qty

                used = transaction.dividend_dps * qty
                found_number_of_shares += qty
                # used_nok = used * transaction.amount.nok_exchange_rate
                total -= used
                d = PortfolioDividend(
                    divdate=transaction.date,
                    qty=qty,
                    dividend_dps=Amount(
                        amountdate=transaction.date,
                        value=transaction.dividend_dps,
                        currency=transaction.amount.currency,
                    ),
                    dividend=Amount(
                        amountdate=transaction.date,
                        value=used,
                        currency=transaction.amount.currency,
                    ),
                    parent=p,
                )
                p.records.append(d)
                if shares_left == 0:
                    break
        if abs(total) > 1:
            logger.error(f"Dividend issue: {transaction.date} Not all dividend used: ${total} expected {expected_number_of_shares}, found: {found_number_of_shares}")
        self.cash.debit(transaction.date, transaction.amount, "dividend")

    def tax(self, transaction):
        """
        TAX type=<EntryTypeEnum.TAX: 'TAX'> date=datetime.date(2022, 10, 26)
        symbol='CSCO' description='Debit' amount=NegativeAmount(currency='USD',
        nok_exchange_rate=Decimal('10.3171'), nok_value=Decimal('-167.446533'),
        value=Decimal('-16.23')) source='schwab:/Users/otroan/Stocks/erik/erik.csv'
        id='TAX 2022-10-26:7'
        """
        self.taxes.append(
            {
                "date": transaction.date,
                "symbol": transaction.symbol,
                "amount": transaction.amount,
            }
        )
        self.cash.credit(transaction.date, transaction.amount, "tax")

    def transfer(self, transaction):
        shares_to_sell = abs(transaction.qty)
        for p in self.positions:
            sold = 0
            if p.symbol == transaction.symbol:
                if p.current_qty == 0:
                    continue
                if p.current_qty >= shares_to_sell:
                    p.current_qty -= shares_to_sell
                    sold = shares_to_sell
                    shares_to_sell = 0
                else:
                    sold = p.current_qty
                    shares_to_sell -= p.current_qty
                    p.current_qty = 0
                s = PortfolioTransfer(
                    date=transaction.date,
                    qty=-sold,
                    parent=p,
                    id=transaction.id,
                )
                p.records.append(s)
                if shares_to_sell == 0:
                    break

    def fee(self, transaction):
        logger.error(f"Fee as a separate record not implemented: {transaction}")

    def cashadjust(self, transaction):
        if transaction.amount.value > 0:
            self.cash.debit(transaction.date, transaction.amount, transaction.description)
        elif transaction.amount.value < 0:
            self.cash.credit(transaction.date, transaction.amount, transaction.description)

    def tax_for_symbol(self, symbol):
        total_nok = 0
        total_usd = 0
        for t in self.taxes:
            if t["symbol"] == symbol:
                total_nok += t["amount"].nok_value
                total_usd += t["amount"].value
        return total_usd, total_nok

    def dividend_reinv(self, transaction: Dividend_Reinv):
        self.cash.credit(transaction.date, transaction.amount, "dividend reinvest")

    def sell(self, transaction):
        shares_to_sell = abs(transaction.qty)

        # This is the net amount after fees
        actual = transaction.amount.value
        # if transaction.fee is not None:
        #     actual -= abs(transaction.fee.value)
        sell_price = Amount(
            amountdate=transaction.date,
            value=(actual / shares_to_sell),
            currency=transaction.amount.currency,
        )
        for p in self.positions:
            sold = 0
            if p.symbol == transaction.symbol:
                if p.current_qty == 0:
                    continue
                if p.current_qty >= shares_to_sell:
                    p.current_qty -= shares_to_sell
                    sold = shares_to_sell
                    shares_to_sell = 0
                else:
                    sold = p.current_qty
                    shares_to_sell -= p.current_qty
                    p.current_qty = 0
                gain_ps = GainAmount.from_amounts(sell_price, p.purchase_price)
                gain = gain_ps * abs(sold)
                s = PortfolioSale(
                    saledate=transaction.date,
                    qty=-sold,
                    sell_price=sell_price,
                    gain_ps=gain_ps,
                    gain=gain,
                    # gain_ps=Amount(amountdate=transaction.date,
                    #                 value=gain_ps.value,
                    #                 currency=transaction.amount.currency),
                    # gain=Amount(amountdate=transaction.date,
                    #                 value=gain.value,
                    #                 currency=transaction.amount.currency),
                    total=sell_price * sold,
                    parent=p,
                    id=transaction.id,
                )
                p.records.append(s)
                if shares_to_sell == 0:
                    break
        self.cash.debit(transaction.date, transaction.amount.model_copy(), "sale")
        # Sale is reported as net value after fees.
        # if transaction.fee is not None:
        #     self.cash.credit(transaction.date, transaction.fee.model_copy(), "sale fee")

    def sell_split(self, transaction, poscopy):
        '''If a sale is split over multiple records, split the position'''
        shares_to_sell = abs(transaction.qty)
        i = 0
        while i < len(poscopy):
            p = poscopy[i]
            if p.symbol != transaction.symbol or p.current_qty == 0:
                i += 1
                continue
            if p.current_qty == shares_to_sell:
                p.current_qty = 0
                shares_to_sell = 0
            elif p.current_qty > shares_to_sell:
                # Split record
                splitpos = deepcopy(p)
                splitpos.current_qty = p.current_qty - shares_to_sell
                splitpos.qty = splitpos.current_qty
                splitpos.pre_split_qty = 0
                splitpos.split = True
                p.current_qty = 0
                self.positions[i].pre_split_qty = p.qty
                self.positions[i].qty = self.positions[i].current_qty = shares_to_sell
                self.positions.insert(i+1, splitpos)
                poscopy.insert(i+1, deepcopy(splitpos))
                logger.debug(f"Splitting position: {p.symbol} {p.date}, {shares_to_sell}+{splitpos.qty}({p.qty})")
                shares_to_sell = 0
            else:
                shares_to_sell -= p.current_qty
                p.current_qty = 0
            if shares_to_sell == 0:
                break
            i += 1

    def buys(self):
        """Return report of BUYS"""
        r = []
        for symbol in self.symbols:
            bought = 0
            price_sum = 0
            price_sum_nok = 0
            no_pos = 0
            for item in self.new_positions:
                if item.symbol != symbol:
                    continue
                bought += item.qty
                price_sum += item.purchase_price.value
                price_sum_nok += item.purchase_price.nok_value
                no_pos += 1
            if no_pos > 0:
                avg_usd = price_sum / no_pos
                avg_nok = price_sum_nok / no_pos
                r.append(
                    {
                        "symbol": symbol,
                        "qty": bought,
                        "avg_usd": avg_usd,
                        "avg_nok": avg_nok,
                    }
                )
        return r

    def sales(self):
        srecords = {}
        for k, v in self.sale_transactions.items():
            srecords[v.id] = EOYSales(
                date=v.date,
                symbol=v.symbol,
                qty=v.qty,
                # fee=r.fee,
                amount=v.amount,
                from_positions=[],
                totals={
                    'gain': NativeAmount(usd_value=0, nok_value=0),
                    'purchase_price': 0,
                    'tax_ded_used': 0,
                    },
            )

        def portfoliosale_to_salesposition(portfoliosale):
            return SalesPosition(
                symbol=portfoliosale.parent.symbol,
                qty=abs(portfoliosale.qty),
                sale_price=portfoliosale.sell_price,
                purchase_price=portfoliosale.parent.purchase_price,
                purchase_date=portfoliosale.parent.date,
                gain_ps=portfoliosale.gain_ps,
                tax_deduction_used=portfoliosale.tax_deduction_used,
            )

        for p in self.positions:
            for r in p.records:
                if isinstance(r, PortfolioSale):
                    record = srecords.get(r.id)
                    record.from_positions.append(portfoliosale_to_salesposition(r))
                    record.totals['gain'] += NativeAmount(
                        usd_value=r.gain.value,
                        nok_value=r.gain.nok_value,
                    )
                    # Average purchase price
                    purchase_price = record.totals['purchase_price']
                    record.totals['purchase_price'] = (purchase_price + r.parent.purchase_price.value) / len(record.from_positions)
                    record.totals['tax_ded_used'] += (r.tax_deduction_used * abs(r.qty))

        sales_report = {}
        for k,v in srecords.items():
            if v.symbol not in sales_report:
                sales_report[v.symbol] = []
            sales_report[v.symbol].append(v)
        return sales_report

    def fees(self):
        # TODO!
        return []

    def wire(self, transaction):
        ''' Handled separately in the cash account '''

    def taxsub(self, transaction):
        self.cash.debit(transaction.date, transaction.amount, "tax returned")

        self.taxes.append(
            {
                "date": transaction.date,
                "symbol": transaction.symbol,
                "amount": transaction.amount,
            }
        )

    dispatch = {
        "BUY": buy,
        "DEPOSIT": buy,
        "SELL": sell,
        "TRANSFER": transfer,
        "DIVIDEND": dividend,
        "DIVIDEND_REINV": dividend_reinv,
        "TAX": tax,
        "TAXSUB": taxsub,
        "WIRE": wire,
        "FEE": fee,
        "CASHADJUST": cashadjust,
    }

    def eoy_balance(self, year):
        """End of year summary of holdings"""
        assert (
            year == self.year or year == self.year - 1
        ), f"Year {year} does not match portfolio year {self.year}"
        end_of_year = f"{year}-12-31"

        eoy_exchange_rate = fmv.get_currency("USD", end_of_year)
        r = []

        if year != self.year:
            if self.prev_holdings and self.prev_holdings.stocks:
                positions = self.prev_holdings.stocks
            else:
                positions = []
        else:
            positions = self.positions
        # positions = self.positions if year == self.year else self.prev_holdings.stocks
        for symbol in self.symbols:
            total_shares = 0
            for p in positions:
                if p.symbol == symbol:
                    try:
                        total_shares += p.current_qty
                    except AttributeError:
                        total_shares += p.qty

            eoyfmv = fmv[symbol, end_of_year]
            r.append(
                EOYBalanceItem(
                    symbol=symbol,
                    qty=total_shares,
                    amount=Amount(
                        value=total_shares * eoyfmv,
                        currency="USD",
                        amountdate=end_of_year,
                        nok_exchange_rate=eoy_exchange_rate,
                        nok_value=total_shares * eoyfmv * eoy_exchange_rate,
                    ),
                    fmv=eoyfmv,
                )
            )
        return r

    def generate_holdings(self, year, broker) -> Holdings:
        # Generate holdings for EOY.
        holdings = []
        assert year == self.year, f"Year {year} does not match portfolio year {self.year}"
        for p in self.positions:
            if p.current_qty == 0:
                continue
            hitem = Stock(
                date=p.date,
                symbol=p.symbol,
                qty=p.current_qty,
                purchase_price=p.purchase_price,
                tax_deduction=p.tax_deduction,
            )
            holdings.append(hitem)

        cash = self.cash_summary.holdings
        return Holdings(year=self.year, broker=broker, stocks=holdings, cash=cash)

    def holdings(self, year, broker):
        return self.generate_holdings(year, broker)

    def fundamentals(self) -> Dict[str, Fundamentals]:
        """Return fundamentals for symbol at date"""
        r = {}
        for symbol in self.symbols:
            fundamentals = fmv.get_fundamentals(symbol)
            isin = fundamentals.get("General", {}).get("ISIN", None)
            if not isin:
                isin = fundamentals.get("ETF_Data", {}).get("ISIN", "")

            r[symbol] = Fundamentals(
                name=fundamentals["General"]["Name"],
                isin=isin,
                country=fundamentals["General"]["CountryName"],
                symbol=fundamentals["General"]["Code"],
            )
        return r

    def dividends(self):
        result = []
        for s in self.symbols:
            # For loop for all self.positions where p.symbol == s
            usd = 0
            nok = 0
            tax_ded_used = 0

            tax_usd, tax_nok = self.tax_for_symbol(s)
            for p in self.positions:
                if p.symbol != s:
                    continue
                for r in p.records:
                    if isinstance(r, PortfolioDividend):
                        usd += r.dividend.value
                        nok += r.dividend.nok_value
                        tax_ded_used += r.tax_deduction_used_total
            result.append(EOYDividend(
                symbol=s,
                amount=NativeAmount(
                    usd_value=usd,
                    nok_value=nok - tax_ded_used,
                ),
                gross_amount=NativeAmount(
                    usd_value=usd,
                    nok_value=nok,
                ),
                tax=NativeAmount(
                    usd_value=tax_usd,
                    nok_value=tax_nok,
                ),
                tax_deduction_used=tax_ded_used,
            ))
        return result

    def synthesize_dividends(self, symbol):
        '''Synthesize dividends for symbol'''
        dividends = fmv.get_dividends(symbol)

        # Filter dividends on self.year
        for k,v in dividends.items():
            try:
                payment_date = todate(k)
                exdate = todate(v['date'])
            except ValueError:
                continue
            if payment_date.year != self.year:
                continue
            print(f'{k}: {v}')
            qty = self.qty_at_date(symbol, exdate)
            print(f'Qty: {qty}')

            if qty > 0:
                amount=PositiveAmount(amountdate=payment_date, currency="USD", value=qty * Decimal(str(v['value'])))
                transaction = Dividend(date=payment_date,
                                       symbol=symbol,
                                       description="Synthesized dividend",
                                       amount=amount,
                                       source='Synthesized dividend')
                self.dividend(transaction)

                tax = Tax(date=payment_date,
                           symbol=symbol,
                           description="Synthesized dividend tax",
                           amount=NegativeAmount(amountdate=payment_date, currency="USD",
                                                 value=Decimal(-0.15) * amount.value),
                           source='Synthesized dividend tax')
                self.tax(tax)

    def espp_extra_info(self):
        """Return extra info for ESPP"""
        r = []
        for p in self.positions:
            if p.discounted_purchase_price:
                r.append({'symbol': p.symbol, 'date': p.date,
                           'discounted_purchase_price_total_nok': p.discounted_purchase_price.nok_value*p.qty,
                           'purchase_price_total_nok': p.purchase_price.nok_value*p.qty})
        return r

    def __init__(  # noqa: C901
        self,
        year: int,
        broker: str,
        transactions: Transactions,
        wires: Wires,
        holdings: Holdings,
        verbose: bool,
        feature_flags: list[FeatureFlagEnum],
    ):
        self.year = year
        self.taxes = []
        self.positions = []
        self.new_positions = []
        if holdings and holdings.cash:
            self.cash = Cash(year=year, opening_balance=holdings.cash)
        else:
            self.cash = Cash(year=year)
        self.broker = broker

        self.column_headers = [
            "Symbol",
            "Date",
            "Type",
            # Qty
            "pQty", # Pre-split qty
            "Qty", # Current qty
            "iQty", # Individual qty after split/sale
            "Price",
            "Price USD",
            "Exchange Rate",

            # Dividends
            "Div PS",
            "Div PS USD",
            "Total Dividend",
            "Total Dividend USD",

            # Deductibe Risk-free return
            "Accumulated",
            "Added",
            "Used",
            "TD Total",

            # Sales
            "Gain PS",
            "Gain PS USD",
            "Gain",
            "Gain USD",
            "Amount",
            "Amount USD",
        ]

        self.prev_holdings = holdings
        if holdings:
            for p in holdings.stocks:
                try:
                    tax_deduction = p.tax_deduction
                except AttributeError:
                    tax_deduction = 0
                self.positions.append(
                    PortfolioPosition(
                        qty=p.qty,
                        purchase_price=p.purchase_price,
                        date=p.date,
                        symbol=p.symbol,
                        tax_deduction_acc=tax_deduction,
                        current_qty=p.qty,
                    )
                )

        #######################
        # Pre-Process buy/sell transactions, split positions as required
        for t in transactions:
            if t.type in ["BUY", "DEPOSIT"]:
                self.buy(t)
        poscopy = deepcopy(self.positions)
        for t in transactions:
            if t.type in ["SELL"]:
                self.sell_split(t, poscopy)
        #######################

        # Dictionary of Sale transactions with transaction id as key
        self.sale_transactions = {t.id: t for t in transactions if t.type == "SELL"}

        for t in transactions:
            # Use dispatch to call a function per t.type
            if t.type in ["BUY", "DEPOSIT"]:    # Already handled these
                continue
            self.__class__.dispatch[t.type](self, t)

        # Add tax deduction to the positions held by the end of the year
        total_tax_deduction = 0

        # Find the set of different symbols in self.positions
        self.symbols = {p.symbol for p in self.positions}

        # Check if we should synthesize dividends
        if FeatureFlagEnum.FEATURE_SYNDIV in feature_flags:
            self.synthesize_dividends(list(self.symbols)[0]) # For now, just use the first symbol

        for p in self.positions:
            if p.current_qty > 0:
                tax_deduction_rate = get_tax_deduction_rate(self.year)
                tax_deduction = (
                    (p.purchase_price.nok_value + p.tax_deduction) * tax_deduction_rate
                ) / 100
                p.tax_deduction_new = tax_deduction
                logger.debug(
                    "Position that gets tax deduction for this year: "
                    f"{p.symbol} {p.date} {p.current_qty} {p.tax_deduction_new}"
                )
            p.tax_deduction = p.tax_deduction_acc + p.tax_deduction_new
            total_tax_deduction += p.tax_deduction * p.qty

        logger.debug(f"Total tax deduction {total_tax_deduction}")

        # Walk through and use the available tax deduction
        # Use tax deduction for dividend if we can. Then for sales. Then keep leftovers for next year.
        for p in self.positions:
            if p.tax_deduction > 0:
                # Walk through records looking for dividends
                for r in p.records:
                    # The qty we get dividends for is different from the qty we get tax dividends for
                    if (
                        isinstance(r, PortfolioDividend)
                        and not FeatureFlagEnum.FEATURE_TFD_ACCUMULATE in feature_flags
                    ):
                        if p.tax_deduction >= r.dividend_dps.nok_value:
                            p.tax_deduction -= r.dividend_dps.nok_value
                            r.tax_deduction_used = r.dividend_dps.nok_value
                            r.tax_deduction_used_total = (
                                r.dividend_dps.nok_value * r.qty
                            )
                        else:
                            r.tax_deduction_used = p.tax_deduction
                            r.tax_deduction_used_total = p.tax_deduction * r.qty
                            p.tax_deduction = 0
                for r in p.records:
                    if isinstance(r, PortfolioSale) and r.gain_ps.nok_value > 0:
                        qty = abs(r.qty)
                        if p.tax_deduction >= r.gain_ps.nok_value:
                            p.tax_deduction -= r.gain_ps.nok_value
                            r.tax_deduction_used = r.gain_ps.nok_value
                            r.tax_deduction_used_total = r.gain_ps.nok_value * qty
                        else:
                            r.tax_deduction_used = p.tax_deduction
                            r.tax_deduction_used_total = p.tax_deduction * qty
                            p.tax_deduction = 0

        # Process wires
        db_wires = [t for t in transactions if t.type == "WIRE"]
        unmatched = self.cash.wire(db_wires, wires)
        self.unmatched_wires = unmatched
        self.unmatched_wires_report = unmatched
        cash_report = self.cash.process()
        self.cash_report = cash_report
        self.cash_summary = cash_report
        # print("CASH REPORT", cash_report)
        self.cash_ledger = self.cash.ledger()
        # print_cash_ledger(self.ledger, console)
        self.ledger = Ledger(holdings, transactions)

        # Generate holdings for next year.
        self.eoy_holdings = self.generate_holdings(year, broker)
        # self.summary = self.generate_tax_summary()

        self.excel_data = self.excel_report()

    def excel_report(self):
        # Create an Excel workbook and get the active sheet
        year = self.year
        portfolio = self.positions
        workbook = Workbook()
        ws = workbook.active
        ws.title = f"Portfolio-{year}"

        disclaimer = ("Disclaimer: This tool is provided as is, without warranty of any kind. "
                    "Use of this tool is at your own risk. The authors or distributors "
                    "are not responsible for any losses, damages, or issues that may arise "
                    "from using this tool. Always consult with a professional financial advisor "
                    "before making any financial decisions."
                    f"This report is generated with the espp2 tool version: {version}")

        # Extract column headers from the Stock Pydantic model
        # Write column headers to the Excel sheet
        ws.merge_cells('J1:M1')
        ws['J1'] = 'Dividends'
        ws['J1'].font = Font(bold=True)
        ws['J1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('N1:Q1')
        ws['N1'] = 'Deductible Risk-free return'
        ws['N1'].font = Font(bold=True)
        ws['N1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('R1:W1')
        ws['R1'] = 'Sales'
        ws['R1'].font = Font(bold=True)
        ws['R1'].alignment = Alignment(horizontal='center', vertical='center')

        ws.append(self.column_headers)
        ft = Font(bold=True)

        title_row = ws.row_dimensions[1]
        title_row.font = ft

        # Write data from Stock instances to the Excel sheet
        row = 3
        for stock in portfolio:
            for row, col, value in stock.format(row, self.column_headers):
                ws.cell(row=row, column=col + 1, value=value)
            row += 1
            for record in stock.records:
                for row, col, value in record.format(row, self.column_headers):
                    ws.cell(row=row, column=col + 1, value=value)
                row += 1

        # Create a dictionary mapping column headers to Excel column letters
        header_to_letter = {header: chr(i + 65) for i, header in enumerate(self.column_headers)}

        # Number format
        num_columns = ["Price", "Price USD", "Gain", "Gain PS",
                       "Gain USD", "Amount", "Amount USD", "Div PS", "Div PS USD",
                       "Total Dividend", "Total Dividend USD", "Exchange Rate",
                       "Accumulated", "Added",]
        num_cols = [header_to_letter[header] for header in num_columns if header in header_to_letter]
        for c in num_cols:
            format_cells(ws, c, "0.00")
        num_columns = ["pQty", "Qty", "iQty",]
        num_cols = [header_to_letter[header] for header in num_columns if header in header_to_letter]
        for c in num_cols:
            format_cells(ws, c, "0.0000")

        # Tax (in a separate sheet?)
        # TODO: Include TAXSUB
        # for t in self.taxes:
        #     ws.append(
        #         [
        #             t["symbol"],
        #             t["date"],
        #             "Tax",
        #             round(t["amount"].nok_value, 2),
        #             round(t["amount"].value, 2),
        #         ]
        #     )

        # Freeze the first row
        c = ws["A2"]
        ws.freeze_panes = c

        adjust_width(ws)
        # Set number format for the entire column
        sum_columns = ["Qty", "Gain", "Gain USD", "Amount", "Amount USD", "Total Dividend",
                    "Total Dividend USD", "TD Total"]

        # Create sum_cols list with Excel column letters
        sum_cols = [header_to_letter[header] for header in sum_columns if header in header_to_letter]

        no_columns = len(ws[sum_cols[0]])
        bold_font = Font(bold=True)
        ws[f"A{no_columns+1}"] = "Total"
        ws[f"A{no_columns+1}"].font = bold_font

        for col in sum_cols:
            ws[f"{col}{no_columns+1}"] = f"=SUM({col}2:{col}{no_columns})"
            ws[f"{col}{no_columns+1}"].font = bold_font
            ws[f"{col}{no_columns+1}"].number_format = "0.00"

        # Format columns with different colors
        format_fill_columns(ws, self.column_headers, ["Div PS", "Div PS USD", "Total Dividend", "Total Dividend USD"], "CAD8EE")
        format_fill_columns(ws, self.column_headers, ["Gain PS", "Gain PS USD", "Gain", "Gain USD", "Amount", "Amount USD"], "90ADD7")
        format_fill_columns(ws, self.column_headers, ["Accumulated", "Added", "TD Total", "Used"], "618CCE")

        # Write the disclaimer to the first cell in the last row
        ws[f"A{ws.max_row + 5}"] = disclaimer

        # Apply conditional formatting to change font color for negative numbers
        ws.conditional_formatting.add(
            ws.dimensions,
            CellIsRule(operator="lessThan", formula=["0.00"], font=Font(color="00FF0000")),
        )

        # Separate sheet for cash
        ws = workbook.create_sheet("Cash")
        ws.append(["Date", "Description", "Amount", "Amount USD", "Total"])
        for c in self.cash_ledger:
            ws.append(
                [
                    c[0].date,
                    c[0].description,
                    round(c[0].amount.nok_value, 2),
                    round(c[0].amount.value, 2),
                    round(c[1], 2),
                ]
            )

        adjust_width(ws)

        # Separate sheet for EOY holdings
        ws = workbook.create_sheet("EOY Holdings")
        ws.append(["Symbol", "Date", "Qty", "Price", "Tax Deduction"])
        for h in self.eoy_holdings.stocks:
            ws.append(
                [
                    h.symbol,
                    h.date,
                    round(h.qty, 4),
                    round(h.purchase_price.nok_value, 2),
                    round(h.tax_deduction, 2),
                ]
            )
        adjust_width(ws)

        # Save the Excel workbook to a binary blob
        excel_data = BytesIO()
        workbook.save(excel_data)
        excel_data.seek(0)
        return excel_data.getvalue()
